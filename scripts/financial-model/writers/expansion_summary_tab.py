from typing import List, Dict, Tuple
from openpyxl import Workbook
from engine.schema import Assumptions
from engine.revenue_year1 import Year1RevenueLine
from engine.revenue_cohort import CohortRevenueLine
from engine.revenue_travel import TravelRevenueLine
from engine.pnl import PnLRow
from writers.styles import HEADER_FILL, HEADER_FONT, CURRENCY_FORMAT, SECTION_FONT, SECTION_FILL, CENTER, LEFT


YEARS = [2026, 2027, 2028, 2029, 2030]


def _revenue_by_year(
    y1_lines: List[Year1RevenueLine],
    cohort_lines: List[CohortRevenueLine],
    travel_lines: List[TravelRevenueLine],
) -> Dict[int, float]:
    totals: Dict[int, float] = {y: 0.0 for y in YEARS}
    for line in y1_lines:
        yr = line.cash_month.year
        if yr in totals:
            totals[yr] += line.net_revenue
    for line in cohort_lines:
        yr = line.cash_month.year
        if yr in totals:
            totals[yr] += line.net_revenue
    for line in travel_lines:
        yr = line.cash_month.year
        if yr in totals:
            totals[yr] += line.net_revenue
    return totals


def _revenue_by_location_product(
    y1_lines: List[Year1RevenueLine],
    cohort_lines: List[CohortRevenueLine],
    travel_lines: List[TravelRevenueLine],
) -> Dict[Tuple[int, str], Dict[int, float]]:
    breakdown: Dict[Tuple[int, str], Dict[int, float]] = {}

    def _add(loc: int, product: str, year: int, amt: float):
        key = (loc, product)
        if key not in breakdown:
            breakdown[key] = {y: 0.0 for y in YEARS}
        if year in breakdown[key]:
            breakdown[key][year] += amt

    for line in y1_lines:
        _add(line.location_id, line.sport, line.cash_month.year, line.net_revenue)
    for line in cohort_lines:
        _add(line.location_id, line.sport, line.cash_month.year, line.net_revenue)
    for line in travel_lines:
        _add(line.location_id, "travel", line.cash_month.year, line.net_revenue)

    return breakdown


def _kids_by_year(
    y1_lines: List[Year1RevenueLine],
    cohort_lines: List[CohortRevenueLine],
    travel_lines: List[TravelRevenueLine],
) -> Dict[int, int]:
    totals: Dict[int, int] = {y: 0 for y in YEARS}
    for line in y1_lines:
        yr = line.cash_month.year
        if yr in totals:
            totals[yr] += line.kids_registered
    for line in cohort_lines:
        yr = line.cash_month.year
        if yr in totals:
            totals[yr] += line.kids_registered
    for line in travel_lines:
        yr = line.cash_month.year
        if yr in totals:
            totals[yr] += line.kids_registered
    return totals


def _ni_by_year(pnl: List[PnLRow]) -> Dict[int, float]:
    totals: Dict[int, float] = {y: 0.0 for y in YEARS}
    for row in pnl:
        yr = row.month.year
        if yr in totals:
            totals[yr] += row.net_income
    return totals


def write_expansion_summary_tab(
    wb: Workbook,
    a: Assumptions,
    y1_lines: List[Year1RevenueLine],
    cohort_lines: List[CohortRevenueLine],
    travel_lines: List[TravelRevenueLine],
    pnl_rows: List[PnLRow],
) -> None:
    ws = wb["Expansion Summary"]
    ws["A1"] = "Expansion Summary — Floor vs. With Expansion"
    ws["A1"].font = SECTION_FONT
    ws["A2"] = "Read this tab first. It shows the single-territory organic baseline (Floor) alongside the full multi-location, multi-sport, travel-enabled story (With Expansion)."

    header_row = 4
    ws.cell(row=header_row, column=1, value="").fill = HEADER_FILL
    for col_idx, year in enumerate(YEARS, start=2):
        c = ws.cell(row=header_row, column=col_idx, value=f"Y{col_idx - 1} ({year})")
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
    c = ws.cell(row=header_row, column=len(YEARS) + 2, value="5yr Total")
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER

    rev_by_year = _revenue_by_year(y1_lines, cohort_lines, travel_lines)
    ni_by_year = _ni_by_year(pnl_rows)
    kids_by_year = _kids_by_year(y1_lines, cohort_lines, travel_lines)

    row = header_row + 1
    ws.cell(row=row, column=1, value="Floor — Single-Territory Rec (Plan 1 baseline)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    floor_refs_ni = {2026: 6054, 2027: 60631, 2028: 84744, 2029: 128660, 2030: 169187}
    ws.cell(row=row, column=1, value="  Reference NI (Plan 1)")
    for col_idx, y in enumerate(YEARS, start=2):
        c = ws.cell(row=row, column=col_idx, value=floor_refs_ni.get(y, 0))
        c.number_format = CURRENCY_FORMAT
    c = ws.cell(row=row, column=len(YEARS) + 2, value=sum(floor_refs_ni.values()))
    c.number_format = CURRENCY_FORMAT
    row += 2

    ws.cell(row=row, column=1, value="With Expansion — Current Model Run").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    ws.cell(row=row, column=1, value="  Total Revenue")
    for col_idx, y in enumerate(YEARS, start=2):
        c = ws.cell(row=row, column=col_idx, value=rev_by_year[y])
        c.number_format = CURRENCY_FORMAT
    c = ws.cell(row=row, column=len(YEARS) + 2, value=sum(rev_by_year.values()))
    c.number_format = CURRENCY_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="  Net Income")
    for col_idx, y in enumerate(YEARS, start=2):
        c = ws.cell(row=row, column=col_idx, value=ni_by_year[y])
        c.number_format = CURRENCY_FORMAT
    c = ws.cell(row=row, column=len(YEARS) + 2, value=sum(ni_by_year.values()))
    c.number_format = CURRENCY_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="  Kids Served")
    for col_idx, y in enumerate(YEARS, start=2):
        ws.cell(row=row, column=col_idx, value=kids_by_year[y])
    ws.cell(row=row, column=len(YEARS) + 2, value=sum(kids_by_year.values()))
    row += 1

    ws.cell(row=row, column=1, value="  Active Locations")
    for col_idx, y in enumerate(YEARS, start=2):
        ws.cell(row=row, column=col_idx, value=a.expansion.locations.by_year.get(y, 1))
    row += 2

    ws.cell(row=row, column=1, value="Breakdown by Location × Product Line").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    breakdown = _revenue_by_location_product(y1_lines, cohort_lines, travel_lines)
    for key in sorted(breakdown.keys()):
        loc, product = key
        label = f"  Loc {loc} — {product}"
        ws.cell(row=row, column=1, value=label).alignment = LEFT
        for col_idx, y in enumerate(YEARS, start=2):
            c = ws.cell(row=row, column=col_idx, value=breakdown[key][y])
            c.number_format = CURRENCY_FORMAT
        total = sum(breakdown[key].values())
        c = ws.cell(row=row, column=len(YEARS) + 2, value=total)
        c.number_format = CURRENCY_FORMAT
        row += 1

    ws.column_dimensions["A"].width = 50
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 16
