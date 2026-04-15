from openpyxl import Workbook
from engine.schema import Assumptions, LowBaseHigh
from writers.styles import (
    HEADER_FILL, HEADER_FONT, SECTION_FILL, SECTION_FONT,
    CONFIDENCE_HIGH, CONFIDENCE_MEDIUM, CONFIDENCE_LOW,
    CURRENCY_FORMAT, PERCENT_FORMAT, CENTER, LEFT,
)


def _write_header(ws, row: int):
    labels = ["Assumption", "Low", "Base", "High", "Unit", "Confidence"]
    for col, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER


def _write_lbh_row(ws, row: int, label: str, lbh: LowBaseHigh, unit: str, confidence: str):
    ws.cell(row=row, column=1, value=label).alignment = LEFT
    ws.cell(row=row, column=2, value=lbh.low)
    ws.cell(row=row, column=3, value=lbh.base)
    ws.cell(row=row, column=4, value=lbh.high)
    ws.cell(row=row, column=5, value=unit)
    fill = {
        "HIGH": CONFIDENCE_HIGH,
        "MEDIUM": CONFIDENCE_MEDIUM,
        "LOW": CONFIDENCE_LOW,
    }.get(confidence, CONFIDENCE_MEDIUM)
    c = ws.cell(row=row, column=6, value=confidence)
    c.fill = fill
    c.alignment = CENTER


def _write_scalar_row(ws, row: int, label: str, value, unit: str):
    ws.cell(row=row, column=1, value=label).alignment = LEFT
    ws.cell(row=row, column=3, value=value)
    ws.cell(row=row, column=5, value=unit)


def write_assumptions_tab(wb: Workbook, a: Assumptions) -> None:
    ws = wb["Assumptions"]
    row = 1
    ws.cell(row=row, column=1, value="PRICING").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1
    _write_header(ws, row)
    row += 1
    _write_lbh_row(ws, row, "Soccer rec price", a.pricing.soccer_price, "$/season", "MEDIUM"); row += 1
    _write_scalar_row(ws, row, "Soccer weeks/season", a.pricing.soccer_weeks_per_season, "weeks"); row += 1
    _write_scalar_row(ws, row, "Soccer seasons/year", a.pricing.soccer_seasons_per_year, "seasons"); row += 1
    _write_scalar_row(ws, row, "Soccer roster size", a.pricing.soccer_roster_size, "kids"); row += 1
    _write_lbh_row(ws, row, "Flag price", a.pricing.flag_price, "$/season", "MEDIUM"); row += 1
    _write_scalar_row(ws, row, "Flag weeks/season", a.pricing.flag_weeks_per_season, "weeks"); row += 1
    _write_scalar_row(ws, row, "Flag seasons/year", a.pricing.flag_seasons_per_year, "seasons"); row += 1
    _write_scalar_row(ws, row, "Flag roster size", a.pricing.flag_roster_size, "kids"); row += 1
    _write_lbh_row(ws, row, "Winter skills $/session", a.pricing.winter_skills_price_per_session, "$/session", "MEDIUM"); row += 1
    _write_scalar_row(ws, row, "Winter skills group size", a.pricing.winter_skills_group_size, "kids"); row += 1
    row += 1

    ws.cell(row=row, column=1, value="DEMAND").font = SECTION_FONT; row += 1
    _write_header(ws, row); row += 1
    _write_lbh_row(ws, row, "Soccer fill rate", a.demand.soccer_fill_rate, "%", "MEDIUM"); row += 1
    _write_lbh_row(ws, row, "Flag fill rate", a.demand.flag_fill_rate, "%", "MEDIUM"); row += 1
    _write_lbh_row(ws, row, "Winter skills fill rate", a.demand.winter_skills_fill_rate, "%", "MEDIUM"); row += 1
    row += 1

    ws.cell(row=row, column=1, value="RETENTION").font = SECTION_FONT; row += 1
    _write_header(ws, row); row += 1
    _write_lbh_row(ws, row, "Soccer S1→S2", a.retention.soccer_s1_to_s2, "%", "MEDIUM"); row += 1
    _write_lbh_row(ws, row, "Soccer S2→S3", a.retention.soccer_s2_to_s3, "%", "MEDIUM"); row += 1
    _write_lbh_row(ws, row, "Soccer S3+", a.retention.soccer_s3_plus, "%", "MEDIUM"); row += 1
    _write_lbh_row(ws, row, "Cross-sell rate", a.retention.cross_sell_rate, "%", "MEDIUM"); row += 1
    _write_lbh_row(ws, row, "Referral multiplier", a.retention.referral_multiplier, "x", "LOW"); row += 1
    row += 1

    ws.cell(row=row, column=1, value="COSTS").font = SECTION_FONT; row += 1
    _write_header(ws, row); row += 1
    _write_lbh_row(ws, row, "Head coach hourly", a.costs.head_coach_hourly, "$/hr", "HIGH"); row += 1
    _write_lbh_row(ws, row, "Assistant coach hourly", a.costs.assistant_coach_hourly, "$/hr", "MEDIUM"); row += 1
    _write_lbh_row(ws, row, "Outdoor field hourly", a.costs.outdoor_field_hourly, "$/hr", "LOW"); row += 1
    _write_lbh_row(ws, row, "Indoor turf full hourly", a.costs.indoor_turf_full_hourly, "$/hr", "MEDIUM"); row += 1
    _write_lbh_row(ws, row, "Indoor turf half hourly", a.costs.indoor_turf_half_hourly, "$/hr", "MEDIUM"); row += 1
    _write_lbh_row(ws, row, "Gym hourly", a.costs.gym_hourly, "$/hr", "HIGH"); row += 1
    _write_scalar_row(ws, row, "Software monthly", a.costs.software_monthly, "$/mo"); row += 1
    _write_scalar_row(ws, row, "Insurance monthly", a.costs.insurance_monthly, "$/mo"); row += 1
    _write_scalar_row(ws, row, "Founder time annual (per founder)", a.costs.founder_time_annual_per_founder, "$/yr"); row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
