from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, size=11)
CONFIDENCE_HIGH = PatternFill("solid", fgColor="C6EFCE")    # green
CONFIDENCE_MEDIUM = PatternFill("solid", fgColor="FFEB9C")  # yellow
CONFIDENCE_LOW = PatternFill("solid", fgColor="FFC7CE")     # red
CURRENCY_FORMAT = '"$"#,##0'
PERCENT_FORMAT = "0.0%"
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
