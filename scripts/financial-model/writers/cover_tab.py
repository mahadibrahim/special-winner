from datetime import date
from openpyxl import Workbook
from engine.schema import Assumptions
from writers.styles import SECTION_FONT


def write_cover_tab(wb: Workbook, a: Assumptions) -> None:
    from openpyxl.styles import Font
    ws = wb["Cover"]
    ws["A1"] = "Aspire Sports — 5-Year Financial Model"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = f"Generated: {date.today().isoformat()}"
    ws["A4"] = f"Horizon: {a.horizon_months} months starting {a.start_month}"
    ws["A5"] = "Source: docs/superpowers/specs/2026-04-15-aspire-sports-financial-model-design.md"

    ws["A7"] = "Confidence legend:"
    ws["A7"].font = SECTION_FONT
    ws["A8"] = "  Green  = HIGH confidence (verified local data)"
    ws["A9"] = "  Yellow = MEDIUM confidence (market research)"
    ws["A10"] = "  Red    = LOW confidence (placeholder, needs outreach)"

    ws["A12"] = "How to use:"
    ws["A12"].font = SECTION_FONT
    ws["A13"] = "  1. Edit values in the Assumptions tab only."
    ws["A14"] = "  2. All downstream tabs reference Assumptions — changes propagate."
    ws["A15"] = "  3. The Scenarios tab shows side-by-side base/downside/upside."
    ws["A16"] = "  4. The Partner Returns tab shows IRR, MOIC, and payback month."

    ws.column_dimensions["A"].width = 80
