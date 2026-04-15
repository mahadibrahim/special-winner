from typing import List
from openpyxl import Workbook
from engine.costs import CostScheduleRow
from writers.styles import HEADER_FILL, HEADER_FONT, CURRENCY_FORMAT, SECTION_FONT, CENTER


def write_costs_tab(wb: Workbook, schedule: List[CostScheduleRow]) -> None:
    ws = wb["Costs"]
    ws["A1"] = "Monthly Cost Schedule"
    ws["A1"].font = SECTION_FONT
    headers = ["Month", "Coach", "Venue", "Uniform", "Software",
               "Insurance", "Bookkeeping", "Founder Time", "Marketing",
               "Curriculum (Expense)", "Curriculum (Cash)", "Total Expense", "Total Cash Out"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER

    for i, r in enumerate(schedule, start=3):
        ws.cell(row=i, column=1, value=r.month.isoformat())
        ws.cell(row=i, column=2, value=r.coach_cost).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=3, value=r.venue_cost).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=4, value=r.uniform_cost).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=5, value=r.software).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=6, value=r.insurance).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=7, value=r.bookkeeping).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=8, value=r.founder_time).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=9, value=r.marketing).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=10, value=r.expense_curriculum).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=11, value=r.cash_curriculum).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=12, value=r.total_expense).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=13, value=r.total_cash_out).number_format = CURRENCY_FORMAT
