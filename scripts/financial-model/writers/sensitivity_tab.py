from typing import List
from openpyxl import Workbook
from engine.sensitivity import TornadoBar
from writers.styles import HEADER_FILL, HEADER_FONT, CURRENCY_FORMAT, SECTION_FONT, CENTER


def write_sensitivity_tab(wb: Workbook, bars: List[TornadoBar]) -> None:
    ws = wb["Sensitivity"]
    ws["A1"] = "One-Variable Sensitivity — Year 3 Net Income"
    ws["A1"].font = SECTION_FONT
    headers = ["Variable", "Low", "Base", "High", "Impact (High − Low)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
    for i, b in enumerate(bars, start=3):
        ws.cell(row=i, column=1, value=b.variable)
        ws.cell(row=i, column=2, value=b.output_low).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=3, value=b.output_base).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=4, value=b.output_high).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=5, value=b.impact).number_format = CURRENCY_FORMAT
    ws.column_dimensions["A"].width = 28
