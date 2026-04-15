from openpyxl import Workbook
from engine.tam import TamReport
from writers.styles import HEADER_FILL, HEADER_FONT, PERCENT_FORMAT, SECTION_FONT, CENTER


def write_tam_tab(wb: Workbook, report: TamReport) -> None:
    ws = wb["TAM Check"]
    ws["A1"] = "TAM Sanity Check — Dublin + Powell"
    ws["A1"].font = SECTION_FONT
    ws["A3"] = "Addressable kids (age 5-12)"
    ws["B3"] = report.addressable_kids
    headers = ["Year", "Implied market share"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
    for i, share in enumerate(report.implied_market_share_by_year, start=6):
        ws.cell(row=i, column=1, value=f"Year {i - 5}")
        c = ws.cell(row=i, column=2, value=share)
        c.number_format = PERCENT_FORMAT
    ws.column_dimensions["A"].width = 26
