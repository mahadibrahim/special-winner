from typing import List
from openpyxl import Workbook
from engine.scenarios import ScenarioResult
from writers.styles import HEADER_FILL, HEADER_FONT, CURRENCY_FORMAT, SECTION_FONT, CENTER


def write_scenarios_tab(wb: Workbook, results: List[ScenarioResult]) -> None:
    ws = wb["Scenarios"]
    ws["A1"] = "Scenarios — Base / Downside / Upside"
    ws["A1"].font = SECTION_FONT

    headers = ["Metric"] + [r.name for r in results]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER

    metrics = [
        ("Year 1 Revenue", "year1_revenue"),
        ("Year 3 Net Income", "year3_net_income"),
        ("Year 5 Net Income", "year5_net_income"),
        ("Min cumulative net income (proxy cash trough)", "min_cash_balance"),
    ]
    for i, (label, attr) in enumerate(metrics, start=4):
        ws.cell(row=i, column=1, value=label)
        for j, r in enumerate(results, start=2):
            c = ws.cell(row=i, column=j, value=getattr(r, attr))
            c.number_format = CURRENCY_FORMAT
    ws.column_dimensions["A"].width = 44
