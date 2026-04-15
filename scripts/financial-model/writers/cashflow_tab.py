from typing import List
from openpyxl import Workbook
from engine.cashflow import CashflowRow
from writers.styles import HEADER_FILL, HEADER_FONT, CURRENCY_FORMAT, SECTION_FONT, CENTER


def write_cashflow_tab(wb: Workbook, rows: List[CashflowRow]) -> None:
    ws = wb["Cash Flow"]
    ws["A1"] = "Monthly Cash Flow"
    ws["A1"].font = SECTION_FONT
    headers = ["Month", "Contributions", "Receipts", "Disbursements", "Net Change", "Ending Balance"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
    for i, r in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=r.month.isoformat())
        ws.cell(row=i, column=2, value=r.contributions).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=3, value=r.receipts).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=4, value=r.disbursements).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=5, value=r.net_change).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=6, value=r.ending_balance).number_format = CURRENCY_FORMAT
