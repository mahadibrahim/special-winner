from typing import List
from openpyxl import Workbook
from engine.pnl import PnLRow
from writers.styles import HEADER_FILL, HEADER_FONT, CURRENCY_FORMAT, SECTION_FONT, CENTER


def write_pnl_tab(wb: Workbook, rows: List[PnLRow]) -> None:
    ws = wb["P&L"]
    ws["A1"] = "Monthly P&L"
    ws["A1"].font = SECTION_FONT
    headers = ["Month", "Revenue", "Variable Cost", "Fixed Expense", "Total Expense", "Net Income"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
    for i, r in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=r.month.isoformat())
        ws.cell(row=i, column=2, value=r.revenue).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=3, value=r.variable_cost).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=4, value=r.fixed_expense).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=5, value=r.total_expense).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=6, value=r.net_income).number_format = CURRENCY_FORMAT
