from typing import List
from openpyxl import Workbook
from engine.revenue_year1 import Year1RevenueLine
from writers.styles import HEADER_FILL, HEADER_FONT, CURRENCY_FORMAT, CENTER, SECTION_FONT


def write_revenue_year1_tab(wb: Workbook, lines: List[Year1RevenueLine]) -> None:
    ws = wb["Revenue Y1"]
    ws["A1"] = "Year 1 Revenue — Bottoms-Up"
    ws["A1"].font = SECTION_FONT

    headers = ["Location", "Sport", "Season", "Year", "Age Band", "Teams", "Kids",
               "Gross Revenue", "Discounts", "Processing", "Net Revenue", "Cash Month"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER

    for i, line in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=line.location_id)
        ws.cell(row=i, column=2, value=line.sport)
        ws.cell(row=i, column=3, value=line.season)
        ws.cell(row=i, column=4, value=line.season_year)
        ws.cell(row=i, column=5, value=line.age_band)
        ws.cell(row=i, column=6, value=line.teams)
        ws.cell(row=i, column=7, value=line.kids_registered)
        gross = ws.cell(row=i, column=8, value=line.gross_revenue); gross.number_format = CURRENCY_FORMAT
        disc = ws.cell(row=i, column=9, value=line.discounts); disc.number_format = CURRENCY_FORMAT
        proc = ws.cell(row=i, column=10, value=line.processing_fees); proc.number_format = CURRENCY_FORMAT
        net = ws.cell(row=i, column=11, value=line.net_revenue); net.number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=12, value=line.cash_month.isoformat())

    for col_letter, width in [("A", 10), ("B", 14), ("C", 10), ("D", 8), ("E", 10),
                              ("F", 8), ("G", 8), ("H", 16), ("I", 12), ("J", 14),
                              ("K", 16), ("L", 14)]:
        ws.column_dimensions[col_letter].width = width
