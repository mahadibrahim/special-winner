from openpyxl import Workbook
from engine.partner_returns import PartnerReturnsReport
from writers.styles import HEADER_FILL, HEADER_FONT, CURRENCY_FORMAT, PERCENT_FORMAT, SECTION_FONT, CENTER


def write_partner_returns_tab(wb: Workbook, report: PartnerReturnsReport) -> None:
    ws = wb["Partner Returns"]
    ws["A1"] = "Partner Returns"
    ws["A1"].font = SECTION_FONT

    # Summary block
    ws["A3"] = "Summary"
    ws["A3"].font = SECTION_FONT
    ws["A4"] = "Total contributions"
    ws["B4"] = sum(report.total_contribution_by_partner.values())
    ws["B4"].number_format = CURRENCY_FORMAT
    ws["A5"] = "Total distributions"
    ws["B5"] = sum(report.total_distribution_by_partner.values())
    ws["B5"].number_format = CURRENCY_FORMAT
    ws["A6"] = "Payback month"
    ws["B6"] = report.payback_month.isoformat() if report.payback_month else "Not reached"
    ws["A7"] = "IRR (annualized)"
    ws["B7"] = report.irr if report.irr is not None else "N/A"
    if report.irr is not None:
        ws["B7"].number_format = PERCENT_FORMAT
    ws["A8"] = "MOIC"
    ws["B8"] = report.moic if report.moic is not None else "N/A"

    # Monthly waterfall
    headers = ["Month", "Contribution", "Business Ending", "Total Distribution"]
    for partner in next(iter(report.monthly_rows)).distribution_by_partner.keys():
        headers.append(f"Dist to {partner}")
        headers.append(f"Cum dist {partner}")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=10, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER

    for i, r in enumerate(report.monthly_rows, start=11):
        ws.cell(row=i, column=1, value=r.month.isoformat())
        ws.cell(row=i, column=2, value=r.contribution).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=3, value=r.business_ending_balance).number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=4, value=r.total_distribution).number_format = CURRENCY_FORMAT
        col = 5
        for p in r.distribution_by_partner:
            ws.cell(row=i, column=col, value=r.distribution_by_partner[p]).number_format = CURRENCY_FORMAT
            ws.cell(row=i, column=col + 1, value=r.cumulative_distribution_by_partner[p]).number_format = CURRENCY_FORMAT
            col += 2
