from typing import List
from openpyxl import Workbook
from engine.revenue_cohort import CohortRevenueLine
from writers.styles import HEADER_FILL, HEADER_FONT, CURRENCY_FORMAT, CENTER, SECTION_FONT


def write_revenue_cohort_tab(wb: Workbook, lines: List[CohortRevenueLine]) -> None:
    ws = wb["Revenue Cohort"]
    ws["A1"] = "Years 2–5 Revenue — Cohort Retention"
    ws["A1"].font = SECTION_FONT

    headers = ["Location", "Sport", "Season", "Year", "Age Band",
               "Kids", "Gross", "Net", "Cash Month"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER

    for i, line in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=line.location_id)
        ws.cell(row=i, column=2, value=line.sport)
        ws.cell(row=i, column=3, value=line.season)
        ws.cell(row=i, column=4, value=line.season_year)
        ws.cell(row=i, column=5, value=line.age_band)
        ws.cell(row=i, column=6, value=line.kids_registered)
        g = ws.cell(row=i, column=7, value=line.gross_revenue); g.number_format = CURRENCY_FORMAT
        n = ws.cell(row=i, column=8, value=line.net_revenue); n.number_format = CURRENCY_FORMAT
        ws.cell(row=i, column=9, value=line.cash_month.isoformat())

    for col_letter, width in [("A", 10), ("B", 14), ("C", 10), ("D", 8), ("E", 10),
                              ("F", 10), ("G", 16), ("H", 16), ("I", 14)]:
        ws.column_dimensions[col_letter].width = width
