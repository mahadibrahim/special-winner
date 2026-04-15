from typing import List, Dict
from openpyxl import Workbook
from engine.schema import Assumptions, SportConfig
from engine.pnl import PnLRow
from engine.costs import (
    CostScheduleRow,
    compute_variable_costs_for_line,
    compute_variable_costs_for_travel_line,
)
from engine.revenue_year1 import Year1RevenueLine
from engine.revenue_cohort import CohortRevenueLine
from engine.revenue_travel import TravelRevenueLine
from engine.calendar import parse_year_month
from writers.styles import HEADER_FILL, HEADER_FONT, CURRENCY_FORMAT, PERCENT_FORMAT, SECTION_FONT, SECTION_FILL, CENTER, LEFT


OP_YEARS = [1, 2, 3, 4, 5]


def _op_year_for_date(d, start_month: str) -> int:
    """Return operational year (1-5) for a given date, or 0 if out of range."""
    start = parse_year_month(start_month)
    months_elapsed = (d.year - start.year) * 12 + (d.month - start.month)
    if months_elapsed < 0:
        return 0
    op_year = (months_elapsed // 12) + 1
    return op_year if 1 <= op_year <= 5 else 0


def write_pnl_annual_tab(
    wb: Workbook,
    a: Assumptions,
    pnl_rows: List[PnLRow],
    cost_schedule: List[CostScheduleRow],
    y1_lines: List[Year1RevenueLine],
    cohort_lines: List[CohortRevenueLine],
    travel_lines: List[TravelRevenueLine],
) -> None:
    """Vertical annual P&L — one row per line item, one column per operational year.

    Line items: Revenue, Coach Payroll, Venue Rental, Uniforms, Total Variable,
    Software, Insurance, Bookkeeping, Marketing, Storage, Founder Time,
    Curriculum Amortization, Total Fixed, Total Expense, Net Income, Margin.
    """
    ws = wb["P&L Annual"]
    ws["A1"] = "Annual P&L — Operational Years (Jul-Jun)"
    ws["A1"].font = SECTION_FONT
    ws["A2"] = "Y1 = Jul 2026 - Jun 2027. Rolls up from monthly engines."

    # Column headers
    header_row = 4
    ws.cell(row=header_row, column=1, value="").fill = HEADER_FILL
    for i, op_year in enumerate(OP_YEARS):
        c = ws.cell(row=header_row, column=1 + i + 1, value=f"Y{op_year}")
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
    c = ws.cell(row=header_row, column=1 + len(OP_YEARS) + 1, value="5yr Total")
    c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER

    # Aggregate variable costs by op year
    sports_by_name: Dict[str, SportConfig] = {s.name: s for s in a.sports}
    coach_by_year = {y: 0.0 for y in OP_YEARS}
    venue_by_year = {y: 0.0 for y in OP_YEARS}
    uniform_by_year = {y: 0.0 for y in OP_YEARS}

    def _add_rec_line(line):
        sport = sports_by_name.get(line.sport)
        if sport is None:
            return
        vc = compute_variable_costs_for_line(a, line, sport)
        oy = _op_year_for_date(line.cash_month, a.start_month)
        if oy:
            coach_by_year[oy] += vc["coach_cost"]
            venue_by_year[oy] += vc["venue_cost"]
            uniform_by_year[oy] += vc["uniform_cost"]

    for l in y1_lines:
        _add_rec_line(l)
    for l in cohort_lines:
        _add_rec_line(l)
    for l in travel_lines:
        vc = compute_variable_costs_for_travel_line(a, l)
        oy = _op_year_for_date(l.cash_month, a.start_month)
        if oy:
            coach_by_year[oy] += vc["coach_cost"]
            venue_by_year[oy] += vc["venue_cost"]
            uniform_by_year[oy] += vc["uniform_cost"]

    # Aggregate fixed costs by op year from the cost schedule
    software_by_year = {y: 0.0 for y in OP_YEARS}
    insurance_by_year = {y: 0.0 for y in OP_YEARS}
    bookkeeping_by_year = {y: 0.0 for y in OP_YEARS}
    marketing_by_year = {y: 0.0 for y in OP_YEARS}
    storage_by_year = {y: 0.0 for y in OP_YEARS}
    founder_by_year = {y: 0.0 for y in OP_YEARS}
    curriculum_by_year = {y: 0.0 for y in OP_YEARS}

    for cs in cost_schedule:
        oy = _op_year_for_date(cs.month, a.start_month)
        if not oy:
            continue
        software_by_year[oy] += cs.software
        insurance_by_year[oy] += cs.insurance
        bookkeeping_by_year[oy] += cs.bookkeeping
        marketing_by_year[oy] += cs.marketing
        storage_by_year[oy] += cs.storage
        founder_by_year[oy] += cs.founder_time
        curriculum_by_year[oy] += cs.expense_curriculum

    # Revenue + NI by op year (from pnl_rows — simpler than re-summing)
    revenue_by_year = {y: 0.0 for y in OP_YEARS}
    ni_by_year = {y: 0.0 for y in OP_YEARS}
    for i, row in enumerate(pnl_rows):
        oy = (i // 12) + 1
        if oy in revenue_by_year:
            revenue_by_year[oy] += row.revenue
            ni_by_year[oy] += row.net_income

    def _write_line(row_num: int, label: str, values: Dict[int, float], bold: bool = False, is_pct: bool = False):
        label_cell = ws.cell(row=row_num, column=1, value=label)
        label_cell.alignment = LEFT
        if bold:
            label_cell.font = SECTION_FONT
        total = 0.0
        for i, oy in enumerate(OP_YEARS):
            v = values[oy]
            total += v
            c = ws.cell(row=row_num, column=1 + i + 1, value=v)
            c.number_format = PERCENT_FORMAT if is_pct else CURRENCY_FORMAT
            if bold:
                c.font = SECTION_FONT
        # 5yr total column
        tc = ws.cell(row=row_num, column=1 + len(OP_YEARS) + 1, value=(total if not is_pct else sum(values.values()) / 5))
        tc.number_format = PERCENT_FORMAT if is_pct else CURRENCY_FORMAT
        if bold:
            tc.font = SECTION_FONT

    # Revenue
    _write_line(5, "Revenue", revenue_by_year, bold=True)

    # Variable cost detail
    ws.cell(row=7, column=1, value="VARIABLE COSTS").font = SECTION_FONT
    _write_line(8, "  Coach payroll", coach_by_year)
    _write_line(9, "  Venue rental", venue_by_year)
    _write_line(10, "  Uniforms / equipment", uniform_by_year)
    total_var_by_year = {oy: coach_by_year[oy] + venue_by_year[oy] + uniform_by_year[oy] for oy in OP_YEARS}
    _write_line(11, "Total Variable", total_var_by_year, bold=True)

    # Fixed cost detail
    ws.cell(row=13, column=1, value="FIXED COSTS").font = SECTION_FONT
    _write_line(14, "  Software & tools", software_by_year)
    _write_line(15, "  Insurance", insurance_by_year)
    _write_line(16, "  Bookkeeping", bookkeeping_by_year)
    _write_line(17, "  Marketing (per-location)", marketing_by_year)
    _write_line(18, "  Storage (HQ)", storage_by_year)
    _write_line(19, "  Founder time imputed", founder_by_year)
    _write_line(20, "  Curriculum amortization", curriculum_by_year)
    total_fixed_by_year = {
        oy: software_by_year[oy] + insurance_by_year[oy] + bookkeeping_by_year[oy]
            + marketing_by_year[oy] + storage_by_year[oy] + founder_by_year[oy]
            + curriculum_by_year[oy]
        for oy in OP_YEARS
    }
    _write_line(21, "Total Fixed", total_fixed_by_year, bold=True)

    # Bottom-line summary
    total_expense_by_year = {oy: total_var_by_year[oy] + total_fixed_by_year[oy] for oy in OP_YEARS}
    _write_line(23, "Total Expense", total_expense_by_year, bold=True)
    _write_line(24, "Net Income", ni_by_year, bold=True)

    # Margin as percentage
    margin_by_year = {
        oy: (ni_by_year[oy] / revenue_by_year[oy] if revenue_by_year[oy] else 0.0)
        for oy in OP_YEARS
    }
    # Write margin separately since we want the 5yr margin to be weighted, not averaged
    label_cell = ws.cell(row=25, column=1, value="Margin %")
    label_cell.alignment = LEFT
    label_cell.font = SECTION_FONT
    for i, oy in enumerate(OP_YEARS):
        c = ws.cell(row=25, column=1 + i + 1, value=margin_by_year[oy])
        c.number_format = PERCENT_FORMAT
        c.font = SECTION_FONT
    total_rev = sum(revenue_by_year.values())
    total_ni = sum(ni_by_year.values())
    total_margin = total_ni / total_rev if total_rev else 0.0
    tc = ws.cell(row=25, column=1 + len(OP_YEARS) + 1, value=total_margin)
    tc.number_format = PERCENT_FORMAT
    tc.font = SECTION_FONT

    # Column widths
    ws.column_dimensions["A"].width = 32
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 14
