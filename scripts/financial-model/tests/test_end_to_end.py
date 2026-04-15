import subprocess
from pathlib import Path
import openpyxl


def test_build_model_generates_xlsx_with_all_tabs(tmp_path):
    root = Path(__file__).parent.parent
    out = root / "output" / "aspire-financial-model.xlsx"
    if out.exists():
        out.unlink()
    result = subprocess.run(
        ["python", "build_model.py"], cwd=root, capture_output=True, text=True
    )
    assert result.returncode == 0, f"build_model.py failed: {result.stderr}"
    assert out.exists()

    wb = openpyxl.load_workbook(out)
    expected_tabs = {
        "Cover", "Assumptions", "Revenue Y1", "Revenue Cohort", "Costs",
        "P&L", "Cash Flow", "Partner Returns", "Sensitivity", "Scenarios", "TAM Check",
    }
    assert expected_tabs.issubset(set(wb.sheetnames))


def test_build_model_pnl_tab_has_monthly_rows():
    root = Path(__file__).parent.parent
    out = root / "output" / "aspire-financial-model.xlsx"
    if not out.exists():
        subprocess.run(["python", "build_model.py"], cwd=root, check=True)
    wb = openpyxl.load_workbook(out)
    ws = wb["P&L"]
    data_rows = [r for r in ws.iter_rows(min_row=3, values_only=True) if r[0] is not None]
    assert len(data_rows) == 60
